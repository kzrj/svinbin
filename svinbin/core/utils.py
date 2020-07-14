# -*- coding: utf-8 -*-
from __future__ import unicode_literals
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side
from datetime import datetime, timedelta

from django.core.exceptions import ValidationError as DjangoValidationError
from rest_framework.exceptions import ValidationError as DRFValidationError

from rest_framework import status, exceptions
from rest_framework.views import exception_handler as drf_exception_handler

from django.utils.encoding import force_text
from django.utils import timezone
from django.core.mail import send_mail
from django.db.utils import IntegrityError as DjangoIntegrityError

from staff.serializers import WorkshopEmployeeSerializer
from locations.models import Location, Section, PigletsGroupCell
from tours.models import MetaTourRecord
from piglets.models import Piglets
from sows_events.models import ( SowFarrow, Semination, Ultrasound, CullingSow, WeaningSow, AbortionSow,
    MarkAsNurse)
from sows.models import SowStatus, Sow, SowStatusRecord
from transactions.models import SowTransaction


class CustomValidation(exceptions.APIException):
    status_code = status.HTTP_400_BAD_REQUEST
    default_detail = 'Error'

    def __init__(self, detail, field, status_code):
        if status_code is not None:
            self.status_code = status_code
        if detail is not None:
            self.detail = {field: force_text(detail)}
        else: self.detail = {'detail': force_text(self.default_detail)}


def custom_exception_handler(exc, context):
    
    if isinstance(exc, CustomValidation):
        field = list(exc.detail.keys())[0]
        exc = DRFValidationError(detail={'message': field + ' ' + exc.detail[field]})

    if isinstance(exc, DjangoValidationError):
        if hasattr(exc, 'message_dict'):
            # TODO: handle many fields
            field = list(exc.detail.keys())[0]
            
            exc = DRFValidationError(detail={'message': field + ' ' + exc.detail[field]})
        else:
            exc = DRFValidationError(detail={'message': exc.message})

    if isinstance(exc, DjangoIntegrityError):
        exc = DRFValidationError(detail={'message': str(exc)})

    return drf_exception_handler(exc, context)


def jwt_response_payload_handler(token, user=None, request=None):
    return {
        'token': token,
        'user': WorkshopEmployeeSerializer(user.employee).data
    }


def export_to_excel_ws3(data, filename='../data/ws3header.xlsx'):
    wb = load_workbook(filename)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]

    row = 7
    col = 1

    Sheet1.cell(1, 24).value = data['results'][0]['date']
    Sheet1.cell(1, 28).value = data['results'][-1]['date']

    for record in data['results']:
        Sheet1.cell(row, col).value = record['date']
        Sheet1.cell(row, col + 1).value = record['count_sows_ws3_start_date']['podsos']
        Sheet1.cell(row, col + 2).value = record['count_sows_ws3_start_date']['suporos']
        Sheet1.cell(row, col + 3).value = record['count_piglets_at_start']
        Sheet1.cell(row, col + 4).value = int(record['count_sows_ws3_start_date']['suporos']) \
            + int(record['count_sows_ws3_start_date']['podsos']) \
            + int(record['count_piglets_at_start'])
        Sheet1.cell(row, col + 5).value = record['tr_in_podsos_count']
        Sheet1.cell(row, col + 6).value = record['tr_in_from_1_sup_count']
        Sheet1.cell(row, col + 7).value = record['tr_in_from_2_sup_count']
        Sheet1.cell(row, col + 9).value = record['count_oporos']
        Sheet1.cell(row, col + 10).value = record['count_alive']
        Sheet1.cell(row, col + 11).value = record['tr_out_podsos_count']
        Sheet1.cell(row, col + 12).value = record['tr_out_sup_count']
        Sheet1.cell(row, col + 13).value = record['padej_podsos_count']
        Sheet1.cell(row, col + 14).value = record['padej_podsos_weight']
        Sheet1.cell(row, col + 15).value = record['padej_sup_count']
        Sheet1.cell(row, col + 16).value = record['padej_sup_weight']
        Sheet1.cell(row, col + 17).value = record['vinuzhd_podsos_count']
        Sheet1.cell(row, col + 18).value = record['vinuzhd_podsos_weight']
        Sheet1.cell(row, col + 19).value = record['vinuzhd_sup_count']
        Sheet1.cell(row, col + 20).value = record['vinuzhd_sup_weight']

        Sheet1.cell(row, col + 21).value = record['tr_out_aka_weight_qnty']
        Sheet1.cell(row, col + 22).value = record['tr_out_aka_weight_total']
        Sheet1.cell(row, col + 23).value = record['tr_out_aka_weight_avg']

        Sheet1.cell(row, col + 24).value = record['piglets_padej_qnty']
        Sheet1.cell(row, col + 25).value = record['piglets_padej_weight']

        Sheet1.cell(row, col + 26).value = record['piglets_vinuzhd_qnty']
        Sheet1.cell(row, col + 27).value = record['piglets_vinuzhd_weight']

        Sheet1.cell(row, col + 28).value = ''
        Sheet1.cell(row, col + 29).value = ''

        Sheet1.cell(row, col + 30).value = record['count_sows_ws3_start_date']['suporos']
        Sheet1.cell(row, col + 31).value = record['count_sows_ws3_start_date']['podsos']
        Sheet1.cell(row, col + 32).value = record['count_piglets_at_start']
        Sheet1.cell(row, col + 33).value = int(record['count_sows_ws3_start_date']['suporos']) \
            + int(record['count_sows_ws3_start_date']['podsos']) \
            + int(record['count_piglets_at_start'])
        row += 1

    Sheet1.cell(row, col).value = 'Итого'
    Sheet1.cell(row, col + 1).value = '-'
    Sheet1.cell(row, col + 2).value = '-'
    Sheet1.cell(row, col + 3).value = '-'
    Sheet1.cell(row, col + 4).value = '-'
    Sheet1.cell(row, col + 5).value = data['total_info']['total_tr_in_podsos_count']
    Sheet1.cell(row, col + 6).value = data['total_info']['total_tr_in_from_1_sup_count']
    Sheet1.cell(row, col + 7).value = data['total_info']['total_tr_in_from_2_sup_count']
    Sheet1.cell(row, col + 9).value = data['total_info']['total_count_oporos']
    Sheet1.cell(row, col + 10).value = data['total_info']['total_count_alive']
    Sheet1.cell(row, col + 11).value = data['total_info']['total_tr_out_podsos_count']
    Sheet1.cell(row, col + 12).value = data['total_info']['total_tr_out_sup_count']
    Sheet1.cell(row, col + 13).value = data['total_info']['total_padej_podsos_count']
    Sheet1.cell(row, col + 14).value = data['total_info']['total_padej_podsos_weight']

    Sheet1.cell(row, col + 15).value = data['total_info']['total_padej_sup_count']
    Sheet1.cell(row, col + 16).value = data['total_info']['total_padej_sup_weight']
    Sheet1.cell(row, col + 17).value = data['total_info']['total_vinuzhd_podsos_count']
    Sheet1.cell(row, col + 18).value = data['total_info']['total_vinuzhd_podsos_weight']
    Sheet1.cell(row, col + 19).value = data['total_info']['total_vinuzhd_sup_count']
    Sheet1.cell(row, col + 20).value = data['total_info']['total_vinuzhd_sup_weight']

    Sheet1.cell(row, col + 21).value = data['total_info']['total_tr_out_aka_weight_qnty']
    Sheet1.cell(row, col + 22).value = data['total_info']['total_tr_out_aka_weight_total']
    Sheet1.cell(row, col + 23).value = data['total_info']['avg_tr_out_weight']
    Sheet1.cell(row, col + 24).value = data['total_info']['total_piglets_padej_qnty']
    Sheet1.cell(row, col + 25).value = data['total_info']['total_piglets_padej_weight']
    Sheet1.cell(row, col + 26).value = data['total_info']['total_piglets_vinuzhd_qnty']
    Sheet1.cell(row, col + 27).value = data['total_info']['total_piglets_vinuzhd_weight']

    Sheet1.cell(row, col + 30).value = '-'
    Sheet1.cell(row, col + 31).value = '-'
    Sheet1.cell(row, col + 32).value = '-'
    Sheet1.cell(row, col + 33).value = '-'

    wb.save("../data/ws3_output.xlsx")


def export_to_excel_ws48(data, ws_number, filename='../data/ws48header.xlsx'):
    wb = load_workbook(filename)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]

    row = 5
    col = 1

    Sheet1.cell(1, 18).value = f'цех {ws_number}'
    Sheet1.cell(1, 23).value = 'с'
    Sheet1.cell(1, 24).value = data['results'][0]['date']
    Sheet1.cell(1, 27).value = 'по'
    Sheet1.cell(1, 28).value = data['results'][-1]['date']

    for idx, record in enumerate(data['results']):
        Sheet1.cell(row, col).value = record['date']
        Sheet1.cell(row, col + 1).value = record['count_piglets_at_start']
        Sheet1.cell(row, col + 2).value = ''
        Sheet1.cell(row, col + 3).value = record['count_piglets_at_start']
        Sheet1.cell(row, col + 4).value = ''

        tr_in_qnty = record['tr_in_qnty'] if record['tr_in_qnty'] else ''
        tr_in_aka_weight_in_qnty = record['tr_in_aka_weight_in_qnty'] if record['tr_in_aka_weight_in_qnty'] else ''
        tr_in_aka_weight_in_qnty = f"({record['tr_in_aka_weight_in_qnty']})" if record['tr_in_aka_weight_in_qnty'] else ''
        Sheet1.cell(row, col + 5).value = f"{tr_in_qnty}{tr_in_aka_weight_in_qnty}"

        Sheet1.cell(row, col + 6).value = record['tr_in_aka_weight_in_total']

        tr_out_qnty = record['tr_out_qnty'] if record['tr_out_qnty'] else ''
        tr_out_aka_weight_in_qnty = f"({record['tr_out_aka_weight_in_qnty']})" if record['tr_out_aka_weight_in_qnty'] else ''
        Sheet1.cell(row, col + 11).value = f"{tr_out_qnty}{tr_out_aka_weight_in_qnty}"

        Sheet1.cell(row, col + 15).value = record['padej_qnty']
        Sheet1.cell(row, col + 16).value = record['padej_total_weight']

        Sheet1.cell(row, col + 25).value = record['vinuzhd_qnty']
        Sheet1.cell(row, col + 26).value = record['vinuzhd_total_weight']

        Sheet1.cell(row, col + 29).value = record['prirezka_qnty']
        Sheet1.cell(row, col + 30).value = record['prirezka_total_weight']

        if idx < len(data['results']) - 1:
            Sheet1.cell(row, col + 31).value = data['results'][idx + 1]['count_piglets_at_start']
            Sheet1.cell(row, col + 32).value = 0
            Sheet1.cell(row, col + 33).value = data['results'][idx + 1]['count_piglets_at_start']

        row += 1

    thin = Side(border_style="thin", color="000000")

    for col_number in range(1, 35):
        Sheet1.cell(row, col_number).font = Font(name='Arial', size=14, bold=True)
        Sheet1.cell(row, col_number).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    Sheet1.cell(row, col).value = 'Итого'
    Sheet1.cell(row, col + 5).value = f"{data['total_info']['total_tr_in_qnty']}({data['total_info']['total_tr_in_aka_weight_in_qnty']})"
    Sheet1.cell(row, col + 6).value = data['total_info']['total_tr_in_aka_weight_in_total']
    Sheet1.cell(row, col + 11).value = f"{data['total_info']['total_tr_out_qnty']}({data['total_info']['total_tr_out_aka_weight_in_qnty']})"

    Sheet1.cell(row, col + 15).value = data['total_info']['total_padej_qnty']
    Sheet1.cell(row, col + 16).value = data['total_info']['total_padej_total_weight']

    Sheet1.cell(row, col + 25).value = data['total_info']['total_vinuzhd_qnty']
    Sheet1.cell(row, col + 26).value = data['total_info']['total_vinuzhd_total_weight']
    
    Sheet1.cell(row, col + 29).value = data['total_info']['total_prirezka_qnty']
    Sheet1.cell(row, col + 30).value = data['total_info']['total_prirezka_total_weight']


    Sheet1.cell(row + 3, col + 11).value = 'Процент падежа от прихода поросят'
    Sheet1.cell(row + 4, col + 11).value = \
        100 * (data['total_info']['total_padej_qnty'] + data['total_info']['total_prirezka_qnty'] 
                + data['total_info']['total_vinuzhd_qnty']) \
        / data['total_info']['total_tr_in_qnty']

    Sheet1.cell(row + 5, col + 11).value = 'Ср.вес 1 головы'
    Sheet1.cell(row + 5, col + 14).value = 'падежа'
    Sheet1.cell(row + 5, col + 15).value = data['total_info']['total_padej_avg_weight']
    Sheet1.cell(row + 5, col + 17).value = 'прирезки'
    Sheet1.cell(row + 5, col + 18).value = data['total_info']['total_prirezka_avg_weight'] 
    Sheet1.cell(row + 5, col + 19).value = 'в.убой'
    Sheet1.cell(row + 5, col + 20).value = data['total_info']['total_vinuzhd_avg_weight'] 

    wb.save(f"../data/ws{ws_number}_output.xlsx")


def export_to_excel_ws567(data, ws_number, filename='../data/ws567header.xlsx'):
    wb = load_workbook(filename)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]

    row = 5
    col = 1

    Sheet1.cell(1, 18).value = f'цех {ws_number}'
    Sheet1.cell(1, 23).value = 'с'
    Sheet1.cell(1, 24).value = data['results'][0]['date']
    Sheet1.cell(1, 27).value = 'по'
    Sheet1.cell(1, 28).value = data['results'][-1]['date']

    for idx, record in enumerate(data['results']):
        Sheet1.cell(row, col).value = record['date']
        Sheet1.cell(row, col + 1).value = record['count_piglets_at_start']
        Sheet1.cell(row, col + 2).value = ''
        Sheet1.cell(row, col + 3).value = record['count_piglets_at_start']
        Sheet1.cell(row, col + 4).value = ''

        tr_in_qnty = record['tr_in_qnty'] if record['tr_in_qnty'] else ''
        tr_in_aka_weight_in_qnty = record['tr_in_aka_weight_in_qnty'] if record['tr_in_aka_weight_in_qnty'] else ''
        tr_in_aka_weight_in_qnty = f"({record['tr_in_aka_weight_in_qnty']})" if record['tr_in_aka_weight_in_qnty'] else ''
        Sheet1.cell(row, col + 5).value = f"{tr_in_qnty}{tr_in_aka_weight_in_qnty}"

        Sheet1.cell(row, col + 6).value = record['tr_in_aka_weight_in_total']

        tr_out_qnty = record['tr_out_qnty'] if record['tr_out_qnty'] else ''
        tr_out_aka_weight_in_qnty = f"({record['tr_out_aka_weight_in_qnty']})" if record['tr_out_aka_weight_in_qnty'] else ''
        Sheet1.cell(row, col + 11).value = f"{tr_out_qnty}{tr_out_aka_weight_in_qnty}"

        Sheet1.cell(row, col + 15).value = record['padej_qnty']
        Sheet1.cell(row, col + 16).value = record['padej_total_weight']

        # Sheet1.cell(row, col + 25).value = record['vinuzhd_qnty']
        # Sheet1.cell(row, col + 26).value = record['vinuzhd_total_weight']

        # Sheet1.cell(row, col + 29).value = record['prirezka_qnty']
        # Sheet1.cell(row, col + 30).value = record['prirezka_total_weight']

        # if idx < len(data['results']) - 1:
        #     Sheet1.cell(row, col + 31).value = data['results'][idx + 1]['count_piglets_at_start']
        #     Sheet1.cell(row, col + 32).value = 0
        #     Sheet1.cell(row, col + 33).value = data['results'][idx + 1]['count_piglets_at_start']

        row += 1

    thin = Side(border_style="thin", color="000000")

    for col_number in range(1, 35):
        Sheet1.cell(row, col_number).font = Font(name='Arial', size=14, bold=True)
        Sheet1.cell(row, col_number).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    Sheet1.cell(row, col).value = 'Итого'
    Sheet1.cell(row, col + 5).value = f"{data['total_info']['total_tr_in_qnty']}({data['total_info']['total_tr_in_aka_weight_in_qnty']})"
    Sheet1.cell(row, col + 6).value = data['total_info']['total_tr_in_aka_weight_in_total']
    Sheet1.cell(row, col + 11).value = f"{data['total_info']['total_tr_out_qnty']}({data['total_info']['total_tr_out_aka_weight_in_qnty']})"

    Sheet1.cell(row, col + 15).value = data['total_info']['total_padej_qnty']
    Sheet1.cell(row, col + 16).value = data['total_info']['total_padej_total_weight']

    Sheet1.cell(row, col + 25).value = data['total_info']['total_vinuzhd_qnty']
    Sheet1.cell(row, col + 26).value = data['total_info']['total_vinuzhd_total_weight']
    
    Sheet1.cell(row, col + 29).value = data['total_info']['total_prirezka_qnty']
    Sheet1.cell(row, col + 30).value = data['total_info']['total_prirezka_total_weight']


    Sheet1.cell(row + 3, col + 11).value = 'Процент падежа от прихода поросят'
    Sheet1.cell(row + 4, col + 11).value = \
        100 * (data['total_info']['total_padej_qnty'] + data['total_info']['total_prirezka_qnty'] 
                + data['total_info']['total_vinuzhd_qnty']) \
        / data['total_info']['total_tr_in_qnty']

    Sheet1.cell(row + 5, col + 11).value = 'Ср.вес 1 головы'
    Sheet1.cell(row + 5, col + 14).value = 'падежа'
    Sheet1.cell(row + 5, col + 15).value = data['total_info']['total_padej_avg_weight']
    Sheet1.cell(row + 5, col + 17).value = 'прирезки'
    Sheet1.cell(row + 5, col + 18).value = data['total_info']['total_prirezka_avg_weight'] 
    Sheet1.cell(row + 5, col + 19).value = 'в.убой'
    Sheet1.cell(row + 5, col + 20).value = data['total_info']['total_vinuzhd_avg_weight'] 

    wb.save(f"../data/ws{ws_number}_output.xlsx")